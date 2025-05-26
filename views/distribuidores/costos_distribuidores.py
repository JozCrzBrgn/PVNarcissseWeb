import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl import load_workbook
from io import BytesIO
import unicodedata
from config.configuration import config
from datetime import datetime as dt

# Leer del estado de sesión
name = st.session_state.get("name")
auth_status = st.session_state.get("auth_status")
username = st.session_state.get("username")

def quitar_acentos(texto):
    return ''.join(
        c for c in unicodedata.normalize('NFKD', texto)
        if not unicodedata.combining(c)
    )

tabla_pedidos_db = {
        "Agrícola Oriental":"db01_pedidos_agri", 
        "Nezahualcóyotl":"db01_pedidos_neza", 
        "Zapotitlán":"db01_pedidos_zapo", 
        "Oaxtepec":"db01_pedidos_oaxt", 
        "Pantitlán":"db01_pedidos_panti",
        "Iztapalapa":"db01_pedidos_iztapalapa",
        "Tonanitla":"db01_pedidos_tona",
        }
tabla_pedidos_celebracion_db = {
        "Agrícola Oriental":"db02_pedidos_celebracion_agri", 
        "Nezahualcóyotl":"db02_pedidos_celebracion_neza", 
        "Zapotitlán":"db02_pedidos_celebracion_zapo", 
        "Oaxtepec":"db02_pedidos_celebracion_oaxt", 
        "Pantitlán":"db02_pedidos_celebracion_panti",
        "Iztapalapa":"db02_pedidos_celebracion_iztapalapa",
        "Tonanitla":"db02_pedidos_celebracion_tona",
        }

if auth_status:
    st.title("Costos Distribuidor")

    sucursal = st.segmented_control(
        "Selecciona un sucursal", 
        ["Agrícola Oriental", "Nezahualcóyotl", "Zapotitlán", "Oaxtepec", "Pantitlán", "Tonanitla", "Iztapalapa"]
        )
    
    if sucursal!=None:
        col1, col2 = st.columns(2)
        with col1:
            hoy = dt.today()
            inicio_mes = hoy.replace(day=1)
            fecha_ini = st.date_input("Selecciona una fecha inicial", value=inicio_mes, key='fi')
        with col2:
            fecha_fin = st.date_input("Selecciona una fecha final", key='ff')


        # Obtener los datos de los productos
        data_prod = config.supabase.table(config.LISTA_PRODUCTOS).select("productos, categoria, precio_ingredientes, precio_compra, precio_venta").execute().data
        # Crear un dataframe
        df_prod = pd.DataFrame(data_prod)
        # Obtener los datos de la tabla
        data = config.supabase.table(tabla_pedidos_db[sucursal]).select("clave, fecha_entrega, producto, cantidad").gte("fecha_entrega", fecha_ini).lte("fecha_entrega", fecha_fin).execute().data
        # Crear un dataframe
        df = pd.DataFrame(data)
        if df.empty:
            st.warning(f"La sucursal de {sucursal} no tiene pedidos de linea en el rango de {fecha_ini} a {fecha_fin}.")
        else:
            #* PIVOTEAMOS POR: clave, producto, PS00..00*n
            pivot_df = df.pivot_table(
                index="producto",
                columns="clave",
                values="cantidad",
                fill_value=0  # reemplaza NaN con 0 si no hay pedido
            )
            # Reset index para dejar "producto" como columna
            pivot_df = pivot_df.reset_index()

            #* Merge por nombre del producto
            df_merge = pivot_df.merge(df_prod,
                                    how="left",
                                    left_on="producto",
                                    right_on="productos")
            # Elimina la columna redundante "productos"
            df_merge.drop(columns=["productos"], inplace=True)
            # Filtrar columnas que comienzan con "PS000"
            cols_ps = [col for col in df_merge.columns if col.startswith("PS000")]

            #* TOTALES
            df_merge["total_piezas"] = df_merge[cols_ps].sum(axis=1)
            # 70% 30%
            df_merge["ingredientes_70"] = df_merge["precio_ingredientes"] * df_merge["total_piezas"] * 0.7
            df_merge["ingredientes_30"] = df_merge["precio_ingredientes"] * df_merge["total_piezas"] * 0.3
            df_merge["compra_70"] = df_merge["precio_compra"] * df_merge["total_piezas"] * 0.7
            df_merge["compra_30"] = df_merge["precio_compra"] * df_merge["total_piezas"] * 0.3
            df_merge["venta_70"] = df_merge["precio_venta"] * df_merge["total_piezas"] * 0.7
            df_merge["venta_30"] = df_merge["precio_venta"] * df_merge["total_piezas"] * 0.3
            # Total precios
            df_merge["total_ingredientes"] = df_merge["precio_ingredientes"] * df_merge["total_piezas"]
            df_merge["total_compra"] = df_merge["precio_compra"] * df_merge["total_piezas"]
            df_merge["total_venta"] = df_merge["precio_venta"] * df_merge["total_piezas"]

            #* PASTELES CELEBRACION
            data_celeb = config.supabase.table(tabla_pedidos_celebracion_db[sucursal]).select("clave, descuento, flete, extras").gte("fecha_entrega", fecha_ini).lte("fecha_entrega", fecha_fin).execute().data
            # Crear un dataframe
            df_celeb = pd.DataFrame(data_celeb)
            if df_celeb.empty:
                df_linea_celeb = df_merge.copy()
            else:
                # Obtenemos la venta total
                df_celeb['total_venta'] = df_celeb['descuento'] + df_celeb['flete'] + df_celeb['extras']
                # Obtenemos 70% y 30%
                df_celeb['venta_70'] = df_celeb['total_venta'] * 0.7
                df_celeb['venta_30'] = df_celeb['total_venta'] * 0.3
                # Eliminamos columnas que no nos sirven
                df_celeb.drop(columns=["descuento", "flete", "extras"], inplace=True)
                # Renombramos columna 'clave' a 'producto'
                df_celeb.rename(columns={'clave': 'producto'}, inplace=True)
                # Agregamos la columna categoria y que todos los renglones digan Celebracion
                df_celeb['categoria'] = 'Celebracion'
                # Crear columnas faltantes en df_celeb con 0 por defecto
                columnas_faltantes = [col for col in df_merge.columns if col not in df_celeb.columns]
                for col in columnas_faltantes:
                    df_celeb[col] = 0
                # Reordenar columnas de df_celeb igual que df_merge
                df_celeb = df_celeb[df_merge.columns]
                # Concatenar los dos DataFrames
                df_linea_celeb = pd.concat([df_merge, df_celeb], ignore_index=True)

            #* ANALISIS DE DATOS
            categorias = df_linea_celeb['categoria'].unique()
            dfs_apilados = []
            for categoria in categorias:
                df_cat = df_linea_celeb[df_linea_celeb['categoria'] == categoria].copy()
                # Calcular la suma de columnas numéricas (excluyendo 'producto' y 'categoria')
                columnas_numericas = df_cat.select_dtypes(include='number').columns
                totales = df_cat[columnas_numericas].sum()
                # Crear fila total
                fila_total = pd.DataFrame(columns=df_cat.columns)
                fila_total.loc[0, 'producto'] = "TOTAL"
                for col in columnas_numericas:
                    fila_total.loc[0, col] = totales[col]
                fila_total['categoria'] = categoria
                # Concatenar con fila de totales
                df_con_totales = pd.concat([df_cat, fila_total], ignore_index=True)
                print(categoria)
                # Agregar título visual
                fila_titulo = pd.DataFrame([[f'=== {categoria.upper()} ==='] + [None]*(df_con_totales.shape[1]-1)], columns=df_con_totales.columns)
                # Agregar todo al acumulador
                dfs_apilados.extend([
                    fila_titulo,
                    df_con_totales,
                    pd.DataFrame([[""] * df_con_totales.shape[1]], columns=df_con_totales.columns)  # fila vacía
                ])
            # Concatenar todo
            df_final = pd.concat(dfs_apilados, ignore_index=True)

            #* GUARDAR EL DATAFRAME EN UN EXCEL CON FORMATO
            # Crear un buffer en memoria
            output = BytesIO()

            # Escribir el DataFrame en el buffer como archivo Excel
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_final.to_excel(writer, index=False, sheet_name="Resumen")

            # Aplicar estilos con openpyxl
            output.seek(0)
            wb = load_workbook(output)
            ws = wb["Resumen"]

            # Estilos
            header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, name="Aptos Narrow")

            total_fill = PatternFill(start_color="C0E6F5", end_color="C0E6F5", fill_type="solid")
            total_font = Font(bold=True, name="Aptos Narrow")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            default_font = Font(name="Aptos Narrow")
            center_align = Alignment(horizontal="center", vertical="center")

            # Encabezado
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align

            # Cuerpo
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = default_font
                    cell.border = thin_border
                    cell.alignment = center_align
                    if row[0].value == "TOTAL":
                        cell.fill = total_fill
                        cell.font = total_font

            # Guardar nuevamente al buffer
            output_formatted = BytesIO()
            wb.save(output_formatted)
            output_formatted.seek(0)

            # Nombre de la sucursal
            nom_sucursal_sin_acentos = quitar_acentos(sucursal)
            nom_sucursal_minusculas = nom_sucursal_sin_acentos.lower()
            nom_sucursal = nom_sucursal_minusculas.replace(" ", "_")

            # Botón para descargar en Streamlit
            st.download_button(
                label="📥 Descargar Excel formateado",
                data=output_formatted,
                file_name=f"resumen_pedidos_{nom_sucursal}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )